# Instalar dependencia: pip install openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── DATOS ─────────────────────────────────────────────────────────────────────
acts = [
    ("1.1.1.1.1","Acta de constitución",4,[],       "Gestión del proyecto"),
    ("1.1.1.1.2","Registro interesados",6,["1.1.1.1.1"],"Gestión del proyecto"),
    ("1.1.1.1.3","Plan de dirección",12,["1.1.1.1.2"],"Gestión del proyecto"),
    ("1.1.1.1.4","Matriz de riesgos",9,["1.1.1.1.3"], "Gestión del proyecto"),
    ("1.1.1.2.1","Informes quincenales",10,["1.1.1.1.4"],"Gestión del proyecto"),
    ("1.1.1.2.2","Indicadores (KPI)",8,["1.1.1.2.1"], "Gestión del proyecto"),
    ("1.1.1.2.3","Reportes de control",7,["1.1.1.2.2"],"Gestión del proyecto"),
    ("1.1.1.3.1","Acta de cierre",5,["1.1.1.2.3"],   "Gestión del proyecto"),
    ("1.1.1.3.2","Lecciones aprendidas",6,["1.1.1.3.1"],"Gestión del proyecto"),
    ("1.1.1.3.3","Dossier transferencia",6,["1.1.1.3.2"],"Gestión del proyecto"),
    ("1.2.1.1.1","Arquitectura sistema",12,[],        "Sistema informático"),
    ("1.2.1.1.2","Módulos funcionales",18,["1.2.1.1.1"],"Sistema informático"),
    ("1.2.1.1.3","Modelo base datos",12,["1.2.1.1.2"],"Sistema informático"),
    ("1.2.1.2.1","Módulo inventario",25,["1.2.1.1.3"],"Sistema informático"),
    ("1.2.1.2.2","Módulo compras",20,["1.2.1.2.1"],  "Sistema informático"),
    ("1.2.1.2.3","Módulo distribución",20,["1.2.1.2.2"],"Sistema informático"),
    ("1.2.1.2.4","Desarrollo físico BD",15,["1.2.1.2.3"],"Sistema informático"),
    ("1.2.1.3.1","Dashboards Power BI",12,["1.2.1.2.4"],"Sistema informático"),
    ("1.2.1.3.2","Alertas tempranas",10,["1.2.1.3.1"],"Sistema informático"),
    ("1.2.1.4.1","Pruebas funcionales",18,["1.2.1.3.2"],"Sistema informático"),
    ("1.2.1.4.2","Implementación nube",12,["1.2.1.4.1"],"Sistema informático"),
    ("1.2.1.4.3","Manual Técnico",10,["1.2.1.4.2"],  "Sistema informático"),
    ("1.3.1.1.1","Manual de funciones",8,[],          "Capital humano"),
    ("1.3.1.1.2","Organigrama",6,["1.3.1.1.1"],      "Capital humano"),
    ("1.3.1.2.1","Plan capacitación",10,["1.3.1.1.2"],"Capital humano"),
    ("1.3.1.2.2","Materiales didácticos",12,["1.3.1.2.1"],"Capital humano"),
    ("1.3.1.2.3","Certificación personal",20,["1.3.1.2.2"],"Capital humano"),
    ("1.4.1.1.1","POE Recepción",6,[],               "Procedimientos operativos"),
    ("1.4.1.1.2","POE Almacenamiento",8,["1.4.1.1.1"],"Procedimientos operativos"),
    ("1.4.1.1.3","POE Despacho",8,["1.4.1.1.2"],     "Procedimientos operativos"),
    ("1.4.1.1.4","Guía FIFO/FEFO",6,["1.4.1.1.3"],  "Procedimientos operativos"),
    ("1.4.1.2.1","Formatos de registro",8,["1.4.1.1.4"],"Procedimientos operativos"),
    ("1.4.1.2.2","Protocolos logísticos",12,["1.4.1.2.1","1.4.1.1.3","1.4.1.1.4"],"Procedimientos operativos"),
    ("1.4.1.2.3","Manual de operación",15,["1.4.1.2.2"],"Procedimientos operativos"),
    ("1.4.1.3.1","Diagnóstico bodegas",12,["1.4.1.2.3"],"Procedimientos operativos"),
    ("1.4.1.3.2","Sistema monitoreo",18,["1.4.1.3.1"],"Procedimientos operativos"),
    ("1.4.1.3.3","Optimización bodega",25,["1.4.1.3.2"],"Procedimientos operativos"),
]

by_id = {a[0]: {"id":a[0],"name":a[1],"dur":a[2],"preds":a[3],"group":a[4],
                 "ES":0,"EF":0,"LS":0,"LF":0,"H":0,"critical":False} for a in acts}

# Paso hacia adelante
changed = True
while changed:
    changed = False
    for a in by_id.values():
        es = max((by_id[p]["EF"] for p in a["preds"] if p in by_id), default=0)
        ef = es + a["dur"]
        if a["ES"] != es or a["EF"] != ef:
            a["ES"], a["EF"] = es, ef
            changed = True

max_ef = max(a["EF"] for a in by_id.values())

# Paso hacia atrás
for a in by_id.values():
    a["LF"] = max_ef
    a["LS"] = max_ef - a["dur"]

changed = True
while changed:
    changed = False
    for a in reversed(list(by_id.values())):
        lf = min((by_id[s["id"]]["LS"] for s in by_id.values() if a["id"] in s["preds"]), default=max_ef)
        ls = lf - a["dur"]
        if a["LF"] != lf or a["LS"] != ls:
            a["LF"], a["LS"] = lf, ls
            changed = True

for a in by_id.values():
    a["H"] = a["LS"] - a["ES"]
    a["critical"] = a["H"] == 0

# ── ESTILOS ───────────────────────────────────────────────────────────────────
def side(style="thin", color="BFBFBF"): return Side(style=style, color=color)
tb = Border(left=side(), right=side(), top=side(), bottom=side())
def fl(h): return PatternFill("solid", start_color=h, fgColor=h)
ca = Alignment(horizontal="center", vertical="center", wrap_text=True)
la = Alignment(horizontal="left",   vertical="center", wrap_text=True)

GC = {
    "Gestión del proyecto":      ("1F5C99","DCEAF9","EEF5FD"),
    "Sistema informático":       ("2E7D32","DCF0DD","E8F5E9"),
    "Capital humano":            ("E65100","FDEBD0","FEF3E8"),
    "Procedimientos operativos": ("6A1B9A","EDE7F6","F3EEF8"),
}

# ── HOJA 1: TABLA PDM ─────────────────────────────────────────────────────────
ws = wb.active
ws.title = "PDM - Tabla"
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A4"

ws.merge_cells("A1:L1")
ws["A1"] = "DIAGRAMA DE RED PDM — GESTIÓN DE PROYECTO LOGÍSTICO"
ws["A1"].font = Font(name="Arial", bold=True, size=16, color="FFFFFF")
ws["A1"].fill = fl("1A3A5C")
ws["A1"].alignment = ca
ws.row_dimensions[1].height = 36

ws.merge_cells("A2:L2")
ws["A2"] = f"Duración total: {max_ef} días  |  Actividades: {len(by_id)}  |  Método: PDM"
ws["A2"].font = Font(name="Arial", size=10, color="FFFFFF", italic=True)
ws["A2"].fill = fl("2C5F8A")
ws["A2"].alignment = ca
ws.row_dimensions[2].height = 22

hdrs    = ["Código","Actividad","Grupo","Predecesoras","Dur\n(días)","ES","EF","LS","LF","Holgura","Ruta\nCrítica","Observación"]
col_widths = [16,28,24,22,8,8,8,8,8,9,10,30]
for ci,(h,w) in enumerate(zip(hdrs,col_widths),1):
    c = ws.cell(row=3, column=ci, value=h)
    c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    c.fill = fl("1A3A5C"); c.alignment = ca; c.border = tb
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.row_dimensions[3].height = 42

prev_g = None; row = 4
for a in by_id.values():
    g = a["group"]; dk,lt,ltr = GC[g]
    if g != prev_g:
        ws.merge_cells(f"A{row}:L{row}")
        c = ws.cell(row=row, column=1, value=f"  ▶  {g.upper()}")
        c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill = fl(dk); c.alignment = la; c.border = tb
        ws.row_dimensions[row].height = 20; row += 1; prev_g = g
    ic = a["critical"]
    rf = fl("FFF3F3") if ic else fl(ltr)
    vals = [a["id"],a["name"],g,", ".join(a["preds"]) or "—",a["dur"],
            a["ES"],a["EF"],a["LS"],a["LF"],a["H"],
            "★ SÍ" if ic else "No",
            "Ruta crítica — holgura cero" if ic else ""]
    for ci,v in enumerate(vals,1):
        c = ws.cell(row=row, column=ci, value=v)
        c.fill = rf; c.border = tb
        c.alignment = la if ci in (2,3,4,12) else ca
        if ci==1:   c.font = Font(name="Arial",bold=True,size=9,color=dk)
        elif ci==10: c.font = Font(name="Arial",bold=True,size=10,color=("C62828" if ic else "2E7D32"))
        elif ci==11: c.font = Font(name="Arial",bold=ic,size=10,color=("C62828" if ic else "555555"))
        else:        c.font = Font(name="Arial",size=10)
    ws.row_dimensions[row].height = 18; row += 1

# ── HOJA 2: GANTT ─────────────────────────────────────────────────────────────
wg = wb.create_sheet("Gantt")
wg.sheet_view.showGridLines = False
wg.merge_cells("A1:C1")
wg["A1"] = "DIAGRAMA GANTT — PDM"
wg["A1"].font = Font(name="Arial",bold=True,size=14,color="FFFFFF")
wg["A1"].fill = fl("1A3A5C"); wg["A1"].alignment = ca
wg.row_dimensions[1].height = 30

OC = 4
wg.column_dimensions["A"].width = 14
wg.column_dimensions["B"].width = 26
wg.column_dimensions["C"].width = 7

for d in range(1, max_ef+1):
    col = d+OC-1
    c = wg.cell(row=2, column=col, value=d)
    c.font = Font(name="Arial",size=7,bold=True,color="FFFFFF")
    c.fill = fl("2C5F8A"); c.alignment = ca; c.border = tb
    wg.column_dimensions[get_column_letter(col)].width = 2.2
wg.row_dimensions[2].height = 18

for ci,h in enumerate(["Código","Actividad","Dur"],1):
    c = wg.cell(row=2,column=ci,value=h)
    c.font = Font(name="Arial",bold=True,size=9,color="FFFFFF")
    c.fill = fl("1A3A5C"); c.alignment = ca; c.border = tb

prev_g = None; row = 3
for a in by_id.values():
    g = a["group"]; dk,lt,ltr = GC[g]; ic = a["critical"]
    if g != prev_g:
        wg.merge_cells(f"A{row}:{get_column_letter(max_ef+OC-1)}{row}")
        c = wg.cell(row=row,column=1,value=f"  {g.upper()}")
        c.font = Font(name="Arial",bold=True,size=9,color="FFFFFF")
        c.fill = fl(dk); c.alignment = la
        wg.row_dimensions[row].height = 16; row += 1; prev_g = g
    for ci,v in enumerate([a["id"],a["name"],a["dur"]],1):
        c = wg.cell(row=row,column=ci,value=v)
        c.font = Font(name="Arial",size=8,bold=(ci==1),color=(dk if ci==1 else "000000"))
        c.fill = fl("FFF3F3") if ic else fl(ltr)
        c.border = tb; c.alignment = la if ci==2 else ca
    bar = "E24B4A" if ic else lt
    for d in range(1, max_ef+1):
        col = d+OC-1; c = wg.cell(row=row,column=col)
        c.fill = fl(bar) if a["ES"] < d <= a["EF"] else fl("F5F5F5")
        c.border = Border(left=Side(style="hair",color="DDDDDD"),
                          right=Side(style="hair",color="DDDDDD"))
    wg.row_dimensions[row].height = 14; row += 1

# ── HOJA 3: RESUMEN ───────────────────────────────────────────────────────────
wr = wb.create_sheet("Resumen")
wr.sheet_view.showGridLines = False
wr.merge_cells("A1:H1")
wr["A1"] = "RESUMEN EJECUTIVO — PDM"
wr["A1"].font = Font(name="Arial",bold=True,size=14,color="FFFFFF")
wr["A1"].fill = fl("1A3A5C"); wr["A1"].alignment = ca
wr.row_dimensions[1].height = 30

kpis = [
    ("Duración total",f"{max_ef} días","1F5C99","DCEAF9"),
    ("Total actividades",str(len(by_id)),"2E7D32","DCF0DD"),
    ("En ruta crítica",str(sum(1 for a in by_id.values() if a["critical"])),"C62828","FFCDD2"),
    ("Con holgura",str(sum(1 for a in by_id.values() if not a["critical"])),"E65100","FFE0B2"),
]
for i,(lbl,val,dk,lt) in enumerate(kpis):
    c1 = i*2+1
    wr.merge_cells(start_row=3,start_column=c1,end_row=3,end_column=c1+1)
    wr.merge_cells(start_row=4,start_column=c1,end_row=4,end_column=c1+1)
    cl = wr.cell(row=3,column=c1,value=lbl)
    cl.font = Font(name="Arial",bold=True,size=9,color="FFFFFF")
    cl.fill = fl(dk); cl.alignment = ca; cl.border = tb
    cv = wr.cell(row=4,column=c1,value=val)
    cv.font = Font(name="Arial",bold=True,size=22,color=dk)
    cv.fill = fl(lt); cv.alignment = ca; cv.border = tb
    wr.row_dimensions[3].height = 22; wr.row_dimensions[4].height = 44
    wr.column_dimensions[get_column_letter(c1)].width = 18
    wr.column_dimensions[get_column_letter(c1+1)].width = 18

wr.row_dimensions[5].height = 10
wr.cell(row=6,column=1,value="RESUMEN POR GRUPO").font = Font(name="Arial",bold=True,size=11,color="1A3A5C")
wr.row_dimensions[6].height = 22

for ci,h in enumerate(["Grupo","Actividades","Dur. acumulada","En ruta crítica","Dur. cadena crítica"],1):
    c = wr.cell(row=7,column=ci,value=h)
    c.font = Font(name="Arial",bold=True,size=9,color="FFFFFF")
    c.fill = fl("1A3A5C"); c.alignment = ca; c.border = tb
    wr.column_dimensions[get_column_letter(ci)].width = [28,14,18,16,20][ci-1]
wr.row_dimensions[7].height = 22

for ri,g in enumerate(["Gestión del proyecto","Sistema informático","Capital humano","Procedimientos operativos"]):
    dk,lt,ltr = GC[g]
    ga = [a for a in by_id.values() if a["group"]==g]
    for ci,v in enumerate([g,len(ga),f"{sum(a['dur'] for a in ga)} días",
                            f"{sum(1 for a in ga if a['critical'])} de {len(ga)}",
                            f"{max((a['EF'] for a in ga),default=0)} días"],1):
        c = wr.cell(row=8+ri,column=ci,value=v)
        c.font = Font(name="Arial",size=9); c.fill = fl(ltr); c.border = tb
        c.alignment = la if ci==1 else ca
    wr.row_dimensions[8+ri].height = 18

wr.row_dimensions[13].height = 10
wr.cell(row=14,column=1,value="DETALLE RUTA CRÍTICA").font = Font(name="Arial",bold=True,size=11,color="C62828")
wr.row_dimensions[14].height = 20

for ci,h in enumerate(["#","Código","Actividad","Grupo","Dur","ES","EF","LS","LF","Holgura"],1):
    c = wr.cell(row=15,column=ci,value=h)
    c.font = Font(name="Arial",bold=True,size=9,color="FFFFFF")
    c.fill = fl("C62828"); c.alignment = ca; c.border = tb
wr.row_dimensions[15].height = 18

for ri,a in enumerate(a for a in by_id.values() if a["critical"]):
    for ci,v in enumerate([ri+1,a["id"],a["name"],a["group"],a["dur"],a["ES"],a["EF"],a["LS"],a["LF"],a["H"]],1):
        c = wr.cell(row=16+ri,column=ci,value=v)
        c.font = Font(name="Arial",size=9,bold=(ci<=2),color=("C62828" if ci==10 else "000000"))
        c.fill = fl("FFF3F3"); c.border = tb
        c.alignment = la if ci in (3,4) else ca
    wr.row_dimensions[16+ri].height = 15

wb.save("PDM_Proyecto.xlsx")
print("✅ Archivo creado: PDM_Proyecto.xlsx")